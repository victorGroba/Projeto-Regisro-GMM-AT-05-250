# preenhcer_carta.py
import sqlite3, os, re, glob
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime, date

# === Caminhos (ajustados) ===
CAMINHO_DB = os.path.join("instance", "temperatura.db")

PASTA_PLANILHAS = r"\\192.168.1.10\Acesso Restrito ISO\Área Técnica\Controles do Setor - MEIOS DE CULTURA\Cartas controle\Carta Controle - GMM-AT-483"
PADRAO_ARQUIVO = "Termômetros - MM-05-AT-483"   # prefixo exibido no print

# procura .xlsx/.xlsm começando com o prefixo acima
candidatos = sorted(glob.glob(os.path.join(PASTA_PLANILHAS, PADRAO_ARQUIVO + "*.xls*")))
if not candidatos:
    # fallback: aceita também arquivos “MM-05-AT-483 …”
    candidatos = sorted(glob.glob(os.path.join(PASTA_PLANILHAS, "*MM-05-AT-483*.xls*")))
if not candidatos:
    raise FileNotFoundError(f"Nenhum arquivo encontrado em:\n{PASTA_PLANILHAS}\n(padrões: '{PADRAO_ARQUIVO}*.xls*' ou '*MM-05-AT-483*.xls*')")

# usa o mais recente
candidatos.sort(key=lambda p: os.path.getmtime(p), reverse=True)
CAMINHO_PLANILHA = candidatos[0]
print("Usando:", os.path.basename(CAMINHO_PLANILHA))

def normaliza(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())

def encontra_aba(wb, identificacao: str):
    if identificacao in wb.sheetnames:
        return wb[identificacao]
    id_norm = normaliza(identificacao)
    for nome in wb.sheetnames:
        if id_norm in normaliza(nome):
            return wb[nome]
    return None

def as_date(v):
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    try:
        return pd.to_datetime(v).date()
    except:  # noqa
        return None

# Carrega banco
conn = sqlite3.connect(CAMINHO_DB)
verificacoes = pd.read_sql_query("SELECT * FROM verificacao", conn)
termometros  = pd.read_sql_query("SELECT * FROM termometro", conn)
conn.close()

verificacoes["data"] = pd.to_datetime(verificacoes["data_hora"]).dt.date
verificacoes = verificacoes.merge(
    termometros[["id","identificacao"]],
    left_on="termometro_id", right_on="id", suffixes=("", "_termometro")
)

wb = load_workbook(CAMINHO_PLANILHA, data_only=True)

inseridos = 0
atualizados = 0

for _, r in verificacoes.iterrows():
    data          = r["data"]
    temperatura   = r["temperatura_atual"]
    responsavel   = r.get("responsavel", "")
    identificacao = r["identificacao"]

    ws = encontra_aba(wb, identificacao)
    if ws is None:
        ident_alt = re.sub(r"[^\w]", "-", identificacao)
        ws = encontra_aba(wb, ident_alt)
        if ws is None:
            continue

    linha = 17
    encontrou = False
    while True:
        valB = ws[f"B{linha}"].value
        if valB is None:
            break
        if as_date(valB) == data and ws[f"C{linha}"].value == temperatura:
            if not ws[f"M{linha}"].value:
                ws[f"M{linha}"] = responsavel
                atualizados += 1
            encontrou = True
            break
        linha += 1

    if not encontrou:
        ws[f"B{linha}"] = data
        ws[f"B{linha}"].number_format = "dd/mm/yyyy"
        ws[f"C{linha}"] = temperatura
        ws[f"M{linha}"] = responsavel
        inseridos += 1

wb.save(CAMINHO_PLANILHA)
print(f"Concluído. Inseridos: {inseridos} | Analistas atualizados: {atualizados}")
